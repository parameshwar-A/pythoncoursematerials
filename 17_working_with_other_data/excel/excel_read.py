import openpyxl

# Loading a workbook from a file
myworkbook = openpyxl.load_workbook("name_list.xlsx")

# Selecting a sheet from a workbook
active_sheet = myworkbook.active

# Reading a value of a particular cell
# print(active_sheet.cell(1,1).value)

# Insert a value in particular cell
# active_sheet.cell(1,1,"lilly_potter")
# print(active_sheet.cell(1,1).value)

# See how many cells have data (non-empty)
# print(active_sheet.dimensions)

# Reading value of a range of cells 
#for item in list(active_sheet.iter_cols())[0]:
#    print(item.value)

# Insert empty columns 
# active_sheet.insert_cols(1, 2)
# print(active_sheet.dimensions)

# Insert empty rows 
# active_sheet.insert_rows(1, 2)
# print(active_sheet.dimensions)

# Move a range of cells to certain cols and rows
# active_sheet.move_range(active_sheet.dimensions, 2,2)
# print(active_sheet.dimensions)

# Inserting multiple values in a sheet
# active_sheet.append(["tony"])
# active_sheet.append(["rhodey"])
# active_sheet.append(["friday"])
# 
# print(active_sheet.dimensions)

# Create a sheet
# myworkbook.create_sheet("mysheet2")
# print(myworkbook.sheetnames)

# Saving the changes
# myworkbook.save("modified_name_list.xlsx")

